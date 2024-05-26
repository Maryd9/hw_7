import io
import csv

from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook

def test_check_pdf(test_create_archive):
    with ZipFile('resources/archive.zip', 'r') as zip_ref:
        with zip_ref.open('tmp/example.pdf') as file:
            pdf_data = file.read()
            reader = PdfReader(io.BytesIO(pdf_data))
            page = reader.pages[0]  # Первая страница
            text = page.extract_text()  # Текст из первой страницы
            assert text == "Тестовый PDF файл"

def test_check_xlsx(test_create_archive):
    with ZipFile('resources/archive.zip', 'r') as zip_ref:
        with zip_ref.open('tmp/import_empl_xlsx.xlsx') as file:
            workbook = load_workbook(file)
            sheet = workbook.active
            first_cell_value = sheet.cell(row=1, column=1).value  # Значение первой ячейки (A1)
            assert first_cell_value == 'Внешний идентификатор для импорта'

def test_check_csv(test_create_archive):
    with ZipFile('resources/archive.zip', 'r') as zip_ref:
        with zip_ref.open('tmp/import_ou_csv.csv') as file:
            csv_text = io.StringIO(file.read().decode('utf-8'))
            csv_reader = csv.reader(csv_text)
            first_row = next(csv_reader)
            first_row_as_string = ','.join(first_row)  # Получение первой строки
            assert first_row_as_string == 'Index,Customer Id,First Name,Last Name,Company,City,Country,Phone 1,Phone 2,Email,Subscription Date,Website'